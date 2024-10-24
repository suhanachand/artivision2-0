#https://dontrepeatyourself.org/post/face-recognition-with-python-dlib-and-deep-learning/
#import dlib
import cv2
from picamera2 import Picamera2

import pickle


import time
#import imutils
#import numpy as np

from utils import face_rects
from utils import face_encodings
from utils import nb_of_matches
from code_logging import logging
import subprocess

BATTERY_TEST = 1 #keep this 1 to check the batter runtime
LIBCAM=1
LOG_FLAG = 1

'''
def logging(text):
  if LOG == 1:
     print(text)
'''

file1 = open("power.txt", "w")
file1.close()

'''
#load power wrokbook
from openpyxl import Workbook
from openpyxl import load_workbook as lwb

wb = Workbook()
ws1 = wb.create_sheet("Sheet1")

#wb = lwb('test.xlsx')
sheet = wb['Sheet1']
'''

start_power_up_time = time.time()
new_time = start_power_up_time

#load the encoding dictionary
with open("encodings.pickle","rb") as f:
   name_encodings_dict = pickle.load(f)

if LIBCAM==0:
   cam = cv2.VideoCapture(0)
   cam.set(3, 640) # set video widht
   cam.set(4, 480) # set video height
else:
   picam2 = Picamera2()
   picam2.configure(picam2.create_preview_configuration(main={"format":'XRGB8888', "size": (1640, 1480)}))

   picam2.start()


while True:
    if LIBCAM==0: #usb cam
       logging("inside cam read",LOG_FLAG )
       ret, img =cam.read()
       logging("img size:"+str(img.shape), LOG_FLAG)
       logging("retrun:"+str(ret),LOG_FLAG)
    else:
       #pcie cm camera
       img = picam2.capture_array() #8bit image
       #img = (img * 256).astype('uint16') #convert to 16bit
       img = cv2.flip(img,0)
       logging(img.shape, LOG_FLAG)
       logging(len(cv2.split(img)), LOG_FLAG)
       b, g, r, c = cv2.split(img)
       img = cv2.merge((b,g,r))
       logging(img.shape, LOG_FLAG)
    #to note the power up
    if BATTERY_TEST == 1:
       if ((time.time() - new_time)) > 60:
          #logging((time.time() - new_time), LOG_FLAG)
          #logging(new_time, LOG_FLAG)
          #logging(time.time(), LOG_FLAG)
          #logging(start_power_up_time, LOG_FLAG)
          uptime = (time.time() - start_power_up_time)/60
          voltage = subprocess.check_output("vcgencmd measure_volts core", shell=True)
          #str_time = f'(%.2f )
          file1 = open("power.txt", "a")
          #logging("\nYour Celsius value is {:0.2f}ºC.\n".format(answer) , LOG_FLAG)
          text = ("uptime in mins:{:0.2f}  voltage:{}\n".format(uptime,voltage))
          logging(text, LOG_FLAG)
          file1.write(text)
          file1.close()
          new_time = time.time()

    #check if blurry
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    var = cv2.Laplacian(img, cv2.CV_64F).var()
    print("Variance= "+str(var))

    if var > 10:
       #img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
       encodings = face_encodings(img)
       names = []
       for encoding in encodings:
           counts = {}
           for (name, encodings) in name_encodings_dict.items():
              counts[name] =  nb_of_matches(encodings, encoding)
       
           if all(count == 0 for count in counts.values()):
              name = "Unknown"
           else:
              name =  max(counts, key=counts.get)

           names.append(name)
      
       for rect, name in zip(face_rects(img), names):
           # get the bounding box for each face using the `rect` variable
           x1, y1, x2, y2 = rect.left(), rect.top(), rect.right(), rect.bottom()
           # draw the bounding box of the face along with the name of the person
           cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 3)
           cv2.putText(img, name, (x1, y1+1), 
                   cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

       # show the output image
       cv2.imshow("image", img)
       k = cv2.waitKey(1) & 0xff # Press 'ESC' for exiting video
       if k == 27:
           break# Do a bit of cleanup

logging("\n [INFO] Exiting Program and cleanup stuff", LOG_FLAG)
cam.release()
cv2.destroyAllWindows()

